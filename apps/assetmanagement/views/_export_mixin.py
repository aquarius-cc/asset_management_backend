"""
导出 Excel Mixin

为所有 ViewSet 提供统一的 Excel 导出能力。
子类只需定义 export_columns 配置即可。
"""

from typing import Any

from django.http import HttpResponse
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated

from utils.response_utils import error_response


class ExportExcelMixin:
    """
    Excel 导出 Mixin

    子类定义 export_columns 配置:
    ```python
    export_columns = [
        {"header": "资产编码", "field": "asset_code"},
        {"header": "资产名称", "field": "asset_name"},
        {"header": "当前状态", "field": "asset_current_status", "display_map": ASSET_STATUS_MAP},
    ]
    ```

    导出 URL: /api/v1/{basename}/export/
    """

    export_columns: list[dict[str, Any]] = []
    export_filename: str = "export.xlsx"
    export_sheet_name: str = "数据导出"

    @action(detail=False, methods=["get"], url_path="export", permission_classes=[IsAuthenticated])
    def export_excel(self, request):
        """导出当前列表数据为 Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return error_response(message="缺少 openpyxl 依赖", status_code=500)

        if not self.export_columns:
            return error_response(message="未配置导出列", status_code=500)

        # 获取行级过滤后的 queryset
        queryset = self.get_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.export_sheet_name

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")

        # 写表头
        for col, col_config in enumerate(self.export_columns, 1):
            cell = ws.cell(row=1, column=col, value=col_config["header"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 写数据
        for row_idx, obj in enumerate(queryset, 2):
            for col, col_config in enumerate(self.export_columns, 1):
                field = col_config["field"]
                value = self._get_nested_value(obj, field)
                display_map = col_config.get("display_map")
                if display_map and value in display_map:
                    value = display_map[value]
                ws.cell(row=row_idx, column=col, value=value or "")

        # 自动调整列宽
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{self.export_filename}"'
        wb.save(response)
        return response

    @staticmethod
    def _get_nested_value(obj, field_path: str):
        """支持嵌套字段访问,如 asset_recordcode__asset_code"""
        parts = field_path.split("__")
        value = obj
        for part in parts:
            if value is None:
                return None
            value = getattr(value, part, None)
        return value
